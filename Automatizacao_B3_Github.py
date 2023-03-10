# -*- coding: utf-8 -*-

#----------------------- Importando as bibliotecas que serão utilizadas durante processo ---------

import pandas as pd #O pandas fornece estruturas de dados para armazenar com eficiência grandes conjuntos de dados, bem como uma ampla variedade de ferramentas para manipulação, limpeza e transformação de dados.
import numpy as np #É uma biblioteca amplamente utilizada para computação numérica em Python. Ele fornece um objeto de matriz multidimensional de alto desempenho, bem como uma grande coleção de funções matemáticas e ferramentas para trabalhar com matrizes.
import time #A biblioteca 'time' em Python fornece funções para trabalhar com hora e data.
from datetime import datetime, date #É uma biblioteca para trabalhar com datas e horas. Ele fornece classes para trabalhar com datas, horas e datas e horas.
from selenium.webdriver.common.action_chains import ActionChains #O ActionChains permite realizar uma série de ações que simulam a interação do usuário com a página web
from selenium import webdriver #O webdriver fornece uma API para automatizar as interações com um navegador da web. Ele permite que você escreva scripts que automatizam tarefas como navegar para uma página da Web, preencher formulários, clicar em botões e muito mais. A API do webdriver oferece suporte a vários navegadores, incluindo Chrome, Firefox, Safari e Internet Explorer.
from msedge.selenium_tools import Edge,EdgeOptions #A classe Edge é a classe principal no pacote Selenium para interagir com o Microsoft Edge. Pode usá-lo para criar uma instância do webdriver e executar várias ações em uma página da Web, como navegar para uma URL, localizar elementos na página e interagir com esses elementos. O EdgeOptions pode usá-lo para definir opções como o caminho executável do driver Microsoft Edge, a página inicial do navegador e muito mais.
#from selenium.webdriver.common.keys import Keys #A classe Keys fornece um conjunto de teclas especiais que podem ser usadas para simular entradas de teclado durante a interação com uma página da Web usando o Selenium WebDriver.
from selenium.webdriver.common.by import By #O By é usado para especificar a maneira como você deseja localizar elementos em uma página da web (Xpath,ID, name,...)
from selenium.webdriver.edge.service import Service #A classe Service é usada para iniciar e parar o serviço Microsoft Edge WebDriver, que permite que o Selenium se comunique com o navegador Microsoft Edge
from selenium.webdriver.edge.options import Options #É uma classe usada para definir opções para o serviço Microsoft Edge WebDriver, que permite que o Selenium se comunique com o navegador Microsoft Edge. Essas opções podem ser usadas para configurar vários aspectos do serviço Microsoft Edge WebDriver, como o número da porta usada para se comunicar com o serviço ou o nível de log do serviço.
from selenium.webdriver.support.wait import WebDriverWait #A classe WebDriverWait no pacote selenium em Python fornece uma maneira conveniente de esperar que uma condição seja atendida antes de prosseguir com um script. Quando você executa um script, às vezes pode levar algum tempo para que um elemento apareça na página ou para que uma página carregue completamente. A classe WebDriverWait ajuda você a esperar que uma determinada condição seja atendida antes de prosseguir com o script.
from selenium.webdriver.support import expected_conditions as EC #O módulo Expected_conditions no pacote Selenium em Python fornece um conjunto de condições predefinidas para aguardar em um script usando a classe WebDriverWait. Essas condições permitem que você espere que eventos específicos ocorram, como um elemento se tornar visível, um elemento se tornar clicável, uma página ser carregada, etc
from openpyxl import load_workbook #O método load_workbook na biblioteca openpyxl em Python é usado para carregar uma pasta de trabalho existente do Excel na memória. Este método retorna um objeto Workbook que pode ser usado para acessar e modificar o conteúdo da pasta de trabalho.
import win32com.client as win32 #É um módulo em Python que fornece acesso à API COM (Component Object Model) da Microsoft. O módulo permite que programas Python interajam com aplicativos do Microsoft Office, como Word, Excel e PowerPoint, bem como outros aplicativos baseados em COM.
import xlwings as xs #É uma biblioteca Python que permite interagir com o Microsoft Excel. Com xlwings, você pode ler e gravar dados em pastas de trabalho do Excel, executar macros e chamar funções do Excel a partir do Python. A biblioteca fornece uma maneira simples e conveniente de automatizar tarefas comuns do Excel, sem a necessidade de interoperabilidade COM complexa ou APIs do Excel de baixo nível.
import ast #É um módulo na biblioteca padrão do Python que fornece uma maneira de "parse" e analisar o código-fonte do Python como uma estrutura semelhante a uma árvore. A árvore representa a sintaxe abstrata do código, com cada nó na árvore representando um elemento sintático diferente do código.
import re #É um módulo na biblioteca padrão do Python que fornece suporte para expressões regulares. Uma expressão regular é um padrão que pode ser usado para corresponder, pesquisar e manipular strings.
import itertools # É um módulo Python que fornece uma coleção de funções para trabalhar com iteradores, que são objetos que podem ser iterados (ou seja, repetidos) para produzir uma sequência de valores. O módulo contém uma variedade de funções que podem ser usadas para criar, manipular e combinar iteradores de várias maneiras.
import os #O módulo "os" no Python fornece uma maneira de interagir com o sistema operacional subjacente no qual o Python está sendo executado. Ele oferece várias funções e métodos para tarefas como operações de arquivos e diretórios, gerenciamento de processos, variáveis ​​de ambiente e muito mais. Com o módulo "os", os programas Python podem executar operações específicas da plataforma, tornando-os mais flexíveis e portáteis em diferentes sistemas operacionais.
import glob #É um módulo que fornece uma maneira conveniente de encontrar todos os nomes de caminho correspondentes a um padrão especificado de acordo com as regras usadas pelo shell Unix
import secrets #O módulo "secrets" é um módulo de biblioteca padrão do Python que fornece uma maneira de gerar números aleatórios seguros e gerenciar segredos, como senhas, chaves de API e outros dados confidenciais.
import string #O módulo "string" é um módulo de biblioteca padrão que fornece várias funções e constantes de manipulação de strings.
import sys #É um módulo que fornece acesso a algumas variáveis ​​e funções relacionadas ao interpretador Python e seu ambiente.
import threading #É um módulo que executa várias threads (unidades menores de um programa) simultaneamente em um único programa. Um thread é um fluxo separado de execução dentro de um programa, e os threads compartilham o mesmo espaço de memória e podem se comunicar uns com os outros.
import tkinter as tk #É uma biblioteca Python que fornece um kit de ferramentas de interface gráfica do usuário (GUI) para criar aplicativos de desktop. É uma das bibliotecas mais usadas para criar aplicativos GUI em Python.
from tkinter import ttk #O módulo ttk no Tkinter significa "tk temático" e fornece um conjunto de widgets aprimorados que têm uma aparência mais moderna em comparação com os widgets Tkinter padrão.
import pythoncom #É um módulo em Python que fornece acesso à tecnologia Component Object Model (COM) da Microsoft, que permite que diferentes componentes de software se comuniquem entre si.


#------------------------  Criação da função "monitoramento_automatico" que será executado ao clicar no botão "run" do GUI para rodar o processo -------------

def monitoramento_automatico(d_i,m_i,a_i,d_f,m_f,a_f):
#-----------------------------------------------------------------------------------------------------

#----------------- Importa e roda os outros scripts do Python para processar os arquivos --------------'
    
    texto_funcionando.config(text="O programa está rodando!") #Altera o texto da Label criada na GUI do Tkinter após iniciar o programa
    
    sys.path.insert(1, '*********************') #Insere o caminho do diretório especificado pela string path na lista sys.path no índice 1. Isso significa que, após a inserção, o Python procurará módulos no diretório especificado por path imediatamente após a pesquisa no diretório de trabalho atual (que está no índice 0 em sys.path).
    
    
    try: #tenta importar, caso não tenha arquivo novos, ele vai continuar o processo
        import Retornos_SRO_REGSS
    except:
        pass
    try:#tenta importar, caso não tenha arquivo novos, ele vai continuar o processo
        import Retornos_SRO_REGPA_MOVPREM
    except:
        pass
    try:#tenta importar, caso não tenha arquivo novos, ele vai continuar o processo
        import Retornos_SRO_MOVPREMRESS
    except:
        pass
    try:#tenta importar, caso não tenha arquivo novos, ele vai continuar o processo
        import Retornos_SRO_REGRESS
    except:
        pass
    
   # -----------------------------------------------------
    
    
    
    #------------------ Data para processamento dos retornos da B3 ------------------------------------
    
    data_inic = date(int(a_i),int(m_i),int(d_i)) #Data inicial com os parâmetros fornecidos na Entry do GUI Tkinter
    data_final = date(int(a_f),int(m_f),int(d_f)) #Data final com os parâmetros fornecidos na Entry do GUI Tkinter
     
    ano_inic = data_inic.strftime("%Y") # Pegando o ano da data inicio para colocar no input do site da b3
    ano_final = data_final.strftime("%Y") # Pegando o ano da data final para colocar no input do site da b3
    mes_inic = data_inic.strftime("%m")  # Pegando o mês da data inicio para colocar no input do site da b3
    mes_final = data_final.strftime("%m")  # Pegando o mês da data final para colocar no input do site da b3
    dia_inic = data_inic.strftime("%d")  # Pegando o dia da data inicio para colocar no input do site da b3
    dia_final = data_final.strftime("%d")  # Pegando o dia da data final para colocar no input do site da b3
    
    #-----------------------------------------------------------------
    
    #------------------- Webscrapping - Retirando a resposta do processamento no InsurConnect B3 ------
    
    options = EdgeOptions() # Parâmetro de opções do Microsoft Edge
    options.add_argument("start-maximized") # Para o API iniciar o Edge com a tela maximizada
    options.add_argument("ignore") 
    
    nav=Edge(executable_path = '**********', options=options) #O webdrive e o parâmetro mostrando o caminho que está localizado o mesmo
    
    
    wait = WebDriverWait(nav, 20) #Faz com que webdriver espere 20 segundos para encontrar objeto, depois disso dá Timeout
    actions = ActionChains(nav) #Objeto de ação de comando para por exemplo: clicar, usar o keyboard, ...
    
    senhas_b3_icatu = pd.read_excel('**********',sheet_name='senha_b3_icatu') #planilha com a senha automática gerada para ICATU
    senhas_b3_riogrande = pd.read_excel('**********',sheet_name='senha_b3_riogrande') #planilha com a senha automática gerada para RIOGRANDEPREV
    
    senhas_b3_icatu = senhas_b3_icatu['ICATU'].tolist() #Colocando as senhas da ICATU em listas para pegar a última
    senhas_b3_riogrande = senhas_b3_riogrande['RIOGRANDEPREV'].tolist() #Colocando as senhas da RIOGRANDEPREV em listas para pegar a última
    
    
    website_b3 ="https://insurconnect.b3.com.br/menu-web/ctp/TelaPrincipalCetip21" #Link do site da b3
    participante=["*******","*******"] #Lista com os dois sistemas Icatu e Riogrande para input do "participante" no site
    usuario = "*******" #Meu usuário no sistema B3
    senha= [senhas_b3_icatu[-1],senhas_b3_riogrande[-1]] #Minha senha no sistema B3 (gerado automaticamente e na planilha)
    
    
    lista_icatu_riogrande = [] #criando lista vazia para adicionar as informações de processamento tanto da Icatu e Riogrande para depois concatenar em um único dataframe
    
    while True: #Colocando looping para caso dê erro no Token, e caso dê erro, try and except, ele vai continuar
        try:
            for i,j in zip(participante,senha): #O for é para obter informação tanto da ICATUSEGUROS quanto RIOGRANDEPREV,e  colocar também suas respectivas senhas
                nav.get(website_b3)     #entrando no website, nesse em especifico no site da B3
                
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"iframe#main"))) #"Entrando" no iframe da página para conseguir encontrar os elementos: participante, usuario e participante
                
                wait.until(EC.element_to_be_clickable((By.ID, "e1"))).send_keys(i) #Colocando o "participante" que seria o grupo ICATU ou RIOGRANDEPREV
                wait.until(EC.element_to_be_clickable((By.ID, "e2"))).send_keys(usuario) #Colocando o meu usuario 
                wait.until(EC.element_to_be_clickable((By.ID, "e3"))).send_keys(j) #Minha senha, que é criada automaticamente toda vez que expira, é retirado de uma planilha de senhas
                wait.until(EC.element_to_be_clickable((By.ID, "Entrada"))).click() #clica no botão "Entrada" para entrar no InsurConnect
                
                message = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#login > table > tbody > tr:nth-child(1) > th"))).get_attribute("innerText") #No elemento de barra azul que contém mensagem, colocar o texto apresentado para fazer condicional dependendo do texto
                
    
                if message == "Sua senha expirou, favor informar uma nova senha.": #O if dentro do for é para caso expire a senha, ele vai gerar uma senha nova, colocar essa senha nova em uma planilha, atualizar a senha e entrar
    
                    letters = string.ascii_letters #Coloca todo o alfabeto, tanto minusculo quanto maiusculo em uma variável
                    digits = string.digits #coloca todo os números em uma variável
                    #special_chars = string.punctuation #Seria para pegar todas as pontuações e colocar em uma variável
    
                    alphabet = letters + digits #variável única com todo o alfabeto e numéricos
    
                    pwd_length = 11 #define o tamanho da senha
    
                    pwd = '@' #gera um único caracter especial para senha que será @
                    for i in range(pwd_length): #vai iterar pelo tamanho definido
                      pwd += ''.join(secrets.choice(alphabet)) #para cada caracter desde o 0 até o 11 será pego aleatório da variável com todos alfabetos e numéricos
    
                    print(pwd) #apenas mostra a senha criada
                    pwd = pd.Series(pwd) #transforma a string em uma serie para inputar na planilha de senhas no Excel
                                    
                    if i == "******": #Se for da ICATU a senha gerada será alimentada na planilha da ICATU
                        path = "*******" #caminho do arquivo de senhas da B3
                        book = load_workbook(path) #abrirá a planinha (invisivel)
                        writer = pd.ExcelWriter(path, engine = 'openpyxl') #Cria um escritor com o mecanismo do Openpyxl
                        writer.book = book #implementar esse parâmetro na planilha aberta
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets) #
    
                        pwd.to_excel(writer, sheet_name = 'senha_b3_icatu', startrow = writer.sheets['senha_b3_icatu'].max_row, index = False, header=False) #Colocando a senha no excel (pwd.to_excel) na planilha especifica ('senha_b3_icatu') colocando na linha inicial 
                                                                                                                                                             #escolhe depois da última linha preenchida(senha) e cola, sem index e sem "nome da coluna"
                        writer.save() #Salva essa escrita na planilha
                    else : #Se for da RIOGRANDE a senha gerada será alimentada na planilha da RIOGRANDE
                        path = "*********" #caminho do arquivo de senhas da B3
                        book = load_workbook(path) #abrirá a planinha (invisivel)
                        writer = pd.ExcelWriter(path, engine = 'openpyxl') #Cria um escritor com o mecanismo do Openpyxl
                        writer.book = book #implementar esse parâmetro na planilha aberta
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
                        pwd.to_excel(writer, sheet_name = 'senha_b3_riogrande', startrow = writer.sheets['senha_b3_icatu'].max_row, index = False, header=False) #Colocando a senha no excel (pwd.to_excel) na planilha especifica ('senha_b3_riogrande') colocando na linha inicial 
                                                                                                                                                             #escolhe depois da última linha preenchida(senha) e cola, sem index e sem "nome da coluna"
                        writer.save()#Salva essa escrita na planilha
                    
                    #------------  Atualizando a senha e entrando para o caso de seja expirada ---------------
                    
                    time.sleep(5) # "adormece" o programa durante 5 segundos, ou seja, fica 5 segundos para executar a próxima instrução
                    wait.until(EC.element_to_be_clickable((By.ID, "e3"))).send_keys(j) #senha antiga
                    time.sleep(1) # "adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    wait.until(EC.element_to_be_clickable((By.ID, "e4"))).send_keys(pwd[0]) #senha nova
                    time.sleep(1) # "adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    wait.until(EC.element_to_be_clickable((By.ID, "e5"))).send_keys(pwd[0]) #senha nova
                    time.sleep(1) # "adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    wait.until(EC.element_to_be_clickable((By.ID, "Entrada"))).click() #atualiza a senha e entra
                    try:
                        # ------------ Caso a reposta da barra azul mostrar "senha complexa" e não criou, refaz a senha conforme anteriormente ---------------
                        while True: #Looping para refazer a senha até dar correto
                            if wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#login > table > tbody > tr:nth-child(1) > td"))).get_attribute("innerText") == 'Senha não atende aos requisitos de comprimento ou complexidade.': #Se a senha colocada foi complexa e deu erro, vai recriar a senha
                                letters = string.ascii_letters
                                digits = string.digits
                                          
                                alphabet = letters + digits 
            
                                pwd_length = 11 
            
                                pwd = '@'
                                for i in range(pwd_length):
                                  pwd += ''.join(secrets.choice(alphabet))
            
                                print(pwd)
                                pwd = pd.Series(pwd)
                                                
                                if i == "ICATUSEG":
                                    path = "**********" #caminho do arquivo
                                    book = load_workbook(path)
                                    writer = pd.ExcelWriter(path, engine = 'openpyxl')
                                    writer.book = book
                                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
                                    pwd.to_excel(writer, sheet_name = 'senha_b3_icatu', startrow = writer.sheets['senha_b3_icatu'].max_row, index = False, header=False)
            
                                    writer.save()
                                else :
                                    path = "************" #caminho do arquivo
                                    book = load_workbook(path)
                                    writer = pd.ExcelWriter(path, engine = 'openpyxl')
                                    writer.book = book
                                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
                                    pwd.to_excel(writer, sheet_name = 'senha_b3_riogrande', startrow = writer.sheets['senha_b3_riogrande'].max_row, index = False, header=False)
            
                                    writer.save()
                                time.sleep(5)
                                wait.until(EC.element_to_be_clickable((By.ID, "e3"))).send_keys(j)
                                time.sleep(1)
                                wait.until(EC.element_to_be_clickable((By.ID, "e4"))).send_keys(pwd[0])
                                time.sleep(1)
                                wait.until(EC.element_to_be_clickable((By.ID, "e5"))).send_keys(pwd[0])
                                time.sleep(1)
                                wait.until(EC.element_to_be_clickable((By.ID, "Entrada"))).click()
                            else: # Se a senha alterou de forma correta, vai continuar o programa
                                break
                                
                    except: #Se der algum erro, ou nao encontrar, vai continuar o programa, pois a senha está correta
                        pass
    
                    # ------------------- Após clickar, gera um token que é enviado para o e-mail para conseguir de fato entrar no Insurconnect -----------------------
                    
                    time.sleep(90) # "adormece" o programa durante 90 segundos, ou seja, fica 90 segundos para executar a próxima instrução
                    pythoncom.CoInitialize() #É uma função no módulo pythoncom da linguagem de programação Python que inicializa a biblioteca COM no thread atual.
                    outlook = win32.Dispatch('outlook.application') #Objeto para abrir o outlook (invisivel)
                    mapi = outlook.GetNamespace("MAPI") #Objeto que é o MAPI (é a aba EMAIL) que contém caixa de entrada, itens enviados, ...
                    inbox = mapi.GetDefaultFolder(6) #Objeto da caixa de entrada (o parâmetro 6 é a pasta INBOX)
                    messages = inbox.Items.Restrict("[SenderEmailAddress] = 'atendimento.seguros@b3.com.br'") #Objeto com as mensagens do INBOX filtrados pelo email da b3 para pegar o Token 
                    lista = list(messages) #Colocando no objeto palpável lista todos os email
                    ultimo_email = lista[-1] #Pegando o último email 
                    view = ultimo_email.HTMLBody #Pegar o corpo de email que de fato é onde contém o texto do token
                    token_conditional = re.compile(r'[0-9]{8}') #Criando o padrão regex para o token, identificando pelo regex onde terão 8 dígitos (numéricos 0 - 9)
                    token_n = int(token_conditional.search(view).group()) #Aqui de fato procura na string o padrão acima e colocando em uma variável transformando em inteiro, onde no input tem que ser esse type
                    time.sleep(10) # "adormece" o programa durante 10 segundos, ou seja, fica 10 segundos para executar a próxima instrução
                    
                    wait.until(EC.element_to_be_clickable((By.ID, "e1"))).send_keys(token_n) #Colocando o token no elemento input do site da b3
                    time.sleep(10) #"adormece" o programa durante 10 segundos, ou seja, fica 10 segundos para executar a próxima instrução
                    wait.until(EC.element_to_be_clickable((By.ID, "Entrada"))).click() #Clica e entra 
                    
                    # ----------------- Entra no Insurconnect para pegar os dados de retorno ---------------
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    
                    alert = wait.until(EC.alert_is_present()) #Coloca o alerta que apresenta ao entrar no Insurconnect da B3 em uma variável
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                                    
                    alert.accept() #aceita o alerta para continuar o processo
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[2]/a/img")))).perform() #Clicka no botão esquerdo superior com "3 barras" para abrir Menu
                    
                    time.sleep(3) #"adormece" o programa durante 3 segundos, ou seja, fica 3 segundos para executar a próxima instrução
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/ul[2]/li[3]/a/span[1]")))).perform() #Clicka no "Transf.de Arquivo" no Menu
                    
                    time.sleep(3) #"adormece" o programa durante 3 segundos, ou seja, fica 3 segundos para executar a próxima instrução
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Arquivo")))).perform() #Clicka no "Arquivo" no Transf.de Arquivo
                    
                    time.sleep(3) #"adormece" o programa durante 3 segundos, ou seja, fica 3 segundos para executar a próxima instrução
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Solicitações de Transferência")))).perform() #Clicka no "Solicitações de Transferência" no Arquivo
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"#main"))) #"Entrando" no iframe da página para conseguir encontrar os elementos: para colocar na "Data de Solicitação"
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[1]")))).send_keys(int(dia_inic)).perform() #Inputa no "Data de Solicitação": dia inicial 
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[2]")))).send_keys(int(mes_inic)).perform() #Inputa no "Data de Solicitação": mês inicial 
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[3]")))).send_keys(int(ano_inic)).perform() #Inputa no "Data de Solicitação": ano inicial 
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[4]")))).send_keys(int(dia_final)).perform() #Inputa no "Data de Solicitação": dia final 
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[5]")))).send_keys(int(mes_final)).perform() #Inputa no "Data de Solicitação": mês final 
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[6]")))).send_keys(int(ano_final)).perform() #Inputa no "Data de Solicitação": ano final 
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[3]/td/table/tbody/tr/td/input[1]")))).perform() #Clica no "Pesquisar" para verificar os retornos do range da data inicial até data final
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    
                    table = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#formArea > table > tbody > tr:nth-child(5) > td > table > tbody > tr > td"))).get_attribute("outerHTML") #Salva a tabela em uma variável 
                    data_first_page=pd.read_html(str(table))[0] #Como a tabela está salva como html, lemos a tabela com read_html
                    
                    lista=[] #Cria uma lista vazia para appendar as informações da tabela de retorno do Insurconnect
                             #Pois pode ter mais de uma página, e assim, mais informação
                                  
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    
                    pagina_numero_ant = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute('value') #Pega o valor numérico do número da página inicial (sempre é 1)
                    pagina_numero_ant = int(pagina_numero_ant) #Coloca o valor como inteiro, porque é extraido como texto
                    
                                    
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr[7]/td/form/a[3]")))).perform() #Clicka no botão para próxima página 
                    
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    pagina_numero_dep = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute("value") #Pega o valor numérico do número da página posterior (se tiver mais de uma página vai ser 2,3,4,5, ... Se não tiver 1)
                    pagina_numero_dep = int(pagina_numero_dep) #Coloca o valor como inteiro, porque é extraido como texto
                    
                    count = 2 #cria variável count inicialmente como 2
                    
                    #----------- Processo inicial para verificação de página e pegar todas as informações de forma correta -------
                    
                    if pagina_numero_ant == pagina_numero_dep: #Se a página anterior for igual a página posterior, continuará o programa 
                        pass
                    else: #Caso não seja igual, vai executar o apendamento das informação de todas as páginas
                        while pagina_numero_ant != pagina_numero_dep and count == pagina_numero_dep: #Se a página anterior for diferente da posterior e o count for igual a página posterior fará o processo de pegar informação
                            time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                            table = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#formArea > table > tbody > tr:nth-child(5) > td > table > tbody > tr > td"))).get_attribute("outerHTML")  #Salva a tabela em uma variável 
                            data_outras_paginas = pd.read_html(str(table))[0] #Como a tabela está salva como html, lemos a tabela com read_html
                            lista.append(data_outras_paginas) #Appenda na lista
                            count = count+1 #count +1
                            ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr[7]/td/form/a[3]")))).perform() #clicka de novo na página para prosseguir para página seguinte
                            time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                            pagina_numero_dep = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute("value") #Substituirá o valor da página posterior pois foi passado de página
                            pagina_numero_dep = int(pagina_numero_dep) #Coloca o valor como inteiro, porque é extraido como texto
                                  
                    
                    if not lista: #Se a lista estiver vazia
                        data_outras_page = lista #Os dados das páginas seguintes estão vazias, ou seja, apenas uma página (1)
                        data_concat = data_first_page #O dado geral será apenas a tabela da página 1
                    
                    else: #Se ela não for vazia
                        data_outras_page = pd.concat(lista) #concatena as outras páginas 
                        data_concat = pd.concat((data_first_page, data_outras_page)) #concatena a página (1) com as outras páginas 
                    
                    
                    lista_icatu_riogrande.append(data_concat) #Colocamos essas tabelas de uma vertente (ICATU ou RIOGRANDE) em uma lista para depois colocar as duas vertentes juntas e ficar uma tabela única
                    
                    time.sleep(3)  #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                    
                    nav.get(website_b3) #Retorna para página inicial                
                                                
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                    
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"#frmMenu"))) #"Entrando" no iframe da página para conseguir encontrar os elementos: sair
                    
                    time.sleep(1) #"adormece" o programa durante 1 segundos, ou seja, fica 1 segundos para executar a próxima instrução
                                
                    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#btnSair > strong"))).click() #Clicka no 'sair' para colocar o próxima empresa do grupo (ICATU ou RIOGRANDE) para pegar a informação de todos eles.
        
                    time.sleep(2) #"adormece" o programa durante 2 segundos, ou seja, fica 2 segundos para executar a próxima instrução
                
                                    
                else: #Ele vai executar o mesmo código anterior, esse else é caso não tenha senha expirada ele prossegue no momento da extração da informação, identico ao anterior
                      #Poderia ter criado função e retornado agora
                    time.sleep(90)
                    pythoncom.CoInitialize() 
                    outlook = win32.Dispatch('outlook.application')
                    mapi = outlook.GetNamespace("MAPI")
                    inbox = mapi.GetDefaultFolder(6) #O parâmetro 6 é a pasta INBOX
                    messages = inbox.Items.Restrict("[SenderEmailAddress] = 'atendimento.seguros@b3.com.br'") #filtrando as mensagens do INBOX oara pegar o Token
                    lista = list(messages)
                    ultimo_email = lista[-1]
                    view = ultimo_email.HTMLBody
                    token_conditional = re.compile(r'[0-9]{8}')
                    token_n = int(token_conditional.search(view).group())
                    print(token_n)
                    time.sleep(5)
                    wait.until(EC.element_to_be_clickable((By.ID, "e1"))).send_keys(token_n)
                    time.sleep(5)
                    wait.until(EC.element_to_be_clickable((By.ID, "Entrada"))).click()
                    
                    time.sleep(5)
                    
                    alert = wait.until(EC.alert_is_present())
                    
                    time.sleep(5)
                    
                    alert.accept()
                                   
                    time.sleep(5)
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[2]/a/img")))).perform()
                    time.sleep(3)
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/ul[2]/li[3]/a/span[1]")))).perform()
                    time.sleep(3)
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Arquivo")))).perform()
                    time.sleep(3)
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Solicitações de Transferência")))).perform()
                    
                    time.sleep(2)
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"#main")))
                    
                    time.sleep(2)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[1]")))).send_keys(int(dia_inic)).perform()
                    
                    time.sleep(1)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[2]")))).send_keys(int(mes_inic)).perform()
                    
                    time.sleep(1)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[3]")))).send_keys(int(ano_inic)).perform()
                    
                    time.sleep(2)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[4]")))).send_keys(int(dia_final)).perform()
                    
                    time.sleep(1)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[5]")))).send_keys(int(mes_final)).perform()
                    
                    time.sleep(1)
                    ActionChains(nav).double_click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td/input[6]")))).send_keys(int(ano_final)).perform()
                    
                    time.sleep(1)
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/table/tbody/tr[3]/td/table/tbody/tr/td/input[1]")))).perform()
                    
                    time.sleep(1)
                    
                    table = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#formArea > table > tbody > tr:nth-child(5) > td > table > tbody > tr > td"))).get_attribute("outerHTML")
                    data_first_page=pd.read_html(str(table))[0]
                    
                    lista=[]
                    
                                  
                    time.sleep(2)
                    
                    pagina_numero_ant = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute('value')
                    pagina_numero_ant = int(pagina_numero_ant)
                    
                                    
                    ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr[7]/td/form/a[3]")))).perform()
                    
                    time.sleep(2)
                    pagina_numero_dep = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute("value")
                    pagina_numero_dep = int(pagina_numero_dep)
                    
                    count = 2
                    
                    if pagina_numero_ant == pagina_numero_dep:
                        pass
                    else:
                        while pagina_numero_ant != pagina_numero_dep and count == pagina_numero_dep:
                            time.sleep(3)
                            table = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#formArea > table > tbody > tr:nth-child(5) > td > table > tbody > tr > td"))).get_attribute("outerHTML")
                            data_outras_paginas = pd.read_html(str(table))[0]
                            lista.append(data_outras_paginas)
                            count = count+1
                            ActionChains(nav).click(wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/table/tbody/tr[7]/td/form/a[3]")))).perform()
                            time.sleep(3)
                            pagina_numero_dep = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#paginacao > input[type=text]:nth-child(4)"))).get_attribute("value")
                            pagina_numero_dep = int(pagina_numero_dep)
                   
                   
                    
                    if not lista:
                        data_outras_page = lista
                        data_concat = data_first_page
                    
                    else:
                        data_outras_page = pd.concat(lista)
                        data_concat = pd.concat((data_first_page, data_outras_page))
                    
                    
                    lista_icatu_riogrande.append(data_concat)
                    
                    time.sleep(3)
                    
                    nav.get(website_b3)               
                                                
                    time.sleep(1)
                    
                    wait.until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"#frmMenu")))
                    
                    time.sleep(1)
                                
                    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#btnSair > strong"))).click()
        
                    time.sleep(2)
                
            nav.close() #Caso acabe as empresas do grupo para extração, vai fechar o navegador
            break #Interrompe o looping feito para caso dê erro de token para não ficar full
            
        except: #Caso dê erro 
            continue #Refaz o looping
        
    
             
    
    data_b3_final = pd.concat(lista_icatu_riogrande) #vai juntar os arquivos da icatu e riogrande extraido do Insurconnect
    data_b3_final = data_b3_final.sort_values(by="Data de Recebimento") #vai ordenar pela data de recebimento
    data_b3_final = data_b3_final.drop_duplicates(keep='first') #deduplica mantendo um, é caso acabe extraindo no looping a mesma informação
    data_b3_final = data_b3_final.reset_index(drop=True) #reseta o index
    
    lista_arquivos_novos_b3 = data_b3_final['Arquivo no Micro'].tolist() #utilizado para validar B3 com retorno via email
    
    
    count_row = data_b3_final.shape[0] #Cria variável para contar quantas linhas tem no arquivo dos valores novos retirado no website da b3
                                       #Para depois filtrar o dataframe como um todo já atualizado com essas informações novas
    
    
    data_b3_final = data_b3_final.style.set_properties(**{
        'font-family': 'Arial',
        'font-size': '8pt',
    }) #Mudando as propriedades do dataframe como a fonte da letra e tamanho para "padronizar" junto com o restante da base histórica 
       #que está na planilha no caminho abaixo
    
    path = "****************" #caminho do arquivo para alimentar a base histórica de retornos da b3 via insurconnect
    
    del alert, actions, i, j, nav, options, pagina_numero_ant, pagina_numero_dep, participante, message, lista_icatu_riogrande #deletando arquivos que não serão utilizados no prosseguimento do código
    del senha, senhas_b3_icatu, senhas_b3_riogrande, token_conditional, table, website_b3 #deletando arquivos que não serão utilizados no prosseguimento do código  
    
    #------------------ Preparando respostas antigas antes de salvar as novas para comparação de respostas para verificação de erros novos --------------------
    
    tabela_respostas_antigas = pd.read_excel(path,usecols='M') #Colocando a tabela histórica em um dataframe
    #tabela_respostas_antigas = tabela_respostas_antigas['Retorno'].tolist()
    
    tabela_respostas_antigas_1 = pd.concat([pd.Series(row['Retorno'].split(',')) for _, row in tabela_respostas_antigas.iterrows()]).reset_index(drop=True) #processo de separação das respostas que estão no valor unificado separado por ,
    #Exemplo coluna='Retorno', linha 1 = ['execucao ok','sinistro já cadastrado'] vai separar em duas linhas iguais mudando só a resposta para coluna='Retorno', linha 1 = 'execucao ok', linha 2 = 'sinistro já cadastrado'
    
    tabela_respostas_antigas_1 = tabela_respostas_antigas_1.rename('Resposta') #Alterando o nome da Serie para Resposta
    tabela_respostas_antigas_1 = tabela_respostas_antigas_1.drop_duplicates(keep='first') #Deduplicando os erros para ficar com único
    
    tabela_respostas_antigas_2 = tabela_respostas_antigas_1.apply(lambda x: 'Sinistro já cadastro' if re.search('já cadastrado',x) is not None else #se encontrar 'sinistro já cadastrado' na resposta retorna o valor da célula como 'sinistro já cadastrado'
              ('Campos inválidos' if re.search('Campos inválidos',x) is not None else  #se encontrar 'campo já cadastrado' na resposta retorna o valor da célula como 'campo já cadastrado
              ('Falha na execucao do callable ' if re.search('Falha na execucao do callable ',x) is not None else x))) #se encontrar 'Falha na execucao do callable' na resposta retorna o valor da célula como 'Falha na execucao do callable'
    tabela_respostas_antigas_2 = tabela_respostas_antigas_2.apply(lambda x: 'Campos inválidos' if re.compile(r'\d{8} inválido').search(x) is not None else x) #se encontrar '8 números + inválido' na resposta retorna o valor da célula como 'Campos inválidos'
    tabela_respostas_antigas_2 = tabela_respostas_antigas_2.apply(lambda x: 'Campos inválidos' if re.compile(r'conteúdo " inválido').search(x) is not None else x) #se encontrar 'conteúdo " inválido' na resposta retorna o valor da célula como 'Campos inválidos'
    
    tabela_respostas_antigas_3 = tabela_respostas_antigas_2.apply(lambda x: x.replace("'","") if re.search("'",x) is not None else x) #Retira ' da resposta
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.apply(lambda x: x.replace("[","") if re.search("\[",x) is not None else x) #Retira [ da resposta
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.apply(lambda x: x.replace("]","") if re.search("\]",x) is not None else x) #Retira ] da resposta
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.replace("",'""') #Retira "" da resposta
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.apply(lambda x: x.strip()) #Retira espaços vazios do inicio e fim da resposta
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.apply(lambda x: 'Necessario informar ao menos uma cobertura para o objeto codigo' if re.search('Necessario informar ao menos uma cobertura para o objeto codigo',x) is not None else x) #se encontrar 'Necessario informar ao menos uma cobertura para o objeto codigo' na resposta retorna o valor da célula como 'Necessario informar ao menos uma cobertura para o objeto codigo'
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.apply(lambda x: 'Campo apolice intermediador invalido' if re.search('apolice intermediador',x) is not None else x) #se encontrar 'apolice intermediador' na resposta retorna o valor da célula como 'apolice intermediador'
    
    tabela_respostas_antigas_3 = tabela_respostas_antigas_3.drop_duplicates(keep='first') #Deduplica as respostas iguais para manter único
    
    del tabela_respostas_antigas, tabela_respostas_antigas_1, tabela_respostas_antigas_2 #deletando arquivos que não serão utilizados no prosseguimento do código
    
    #-----------------------------------------------------------------------------
    
    #----------------------------- Salvando as resposta novas (retirada com webscrapping) dentro da planilha de histórico
    
    book = load_workbook(path, keep_vba=True) #abrirá a planinha (invisivel)
    writer = pd.ExcelWriter(path, engine = 'openpyxl') #Cria um escritor com o mecanismo do Openpyxl
    writer.book = book #implementar esse parâmetro na planilha aberta
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    data_b3_final.to_excel(writer, sheet_name = 'Lista arquivos', startrow = writer.sheets['Lista arquivos'].max_row, index = False, header=False) #Salvando os dados novos (data_b3_final.to_excel) na planilha histórica ('Lista arquivos') colocando na linha inicial 
                                                                                                                                         #depois da última linha preenchida(última informação do histórico) e cola, sem index e sem "nome da coluna"
    
    writer.save() #Salva essa escrita na planilha
    
    # ----------------------------------------------------
    
    # ------------------ Importa e roda os outros scripts do Python para processar os arquivos -------------
    
    #Tem que rodar nessa etapa, pois é necessário colocar as informações novas no histórico para funcionar o script, pois tem como referência os arquivos novos
    try: #tentar importar, caso não tenha arquivo novos, ele vai continuar o processo
        import TodosRetornosConsolidados
    except:
        pass
    
    #-----------------------------------------------------
    
    # ---------------- Abertura da planilha para ativar a macro (open_workbook) para modificar formatação e filldown de fórmula ------------------
    xl_app = xs.App(visible=False, add_book=False) #parâmetro aplicativo excel, visible = false, não mostrará e não adicionará em um novo book
    wb = xl_app.books.open(path) #abrir a planilha 
    
    
    wb.save() # salva e
    wb.close() #fecha (apenas para ativar a macro)
    
    xl_app.quit() #sai do aplicativo Excel
    
    #---------------------------------------------
    
    data_inic = data_inic.strftime('%d-%m-%Y') #transforma a variável data_inic, que inputamos no formato string dia - mes - ano
    data_final = data_final.strftime('%d-%m-%Y') #transforma a variável data_final, que inputamos no formato string dia - mes - ano
    
    data_pendente = pd.read_excel(path) #Coloca a tabela completa na variável
    
    data_pendente['Data de Recebimento'] = data_pendente['Data de Recebimento'].apply(lambda x: x.strftime('%d-%m-%Y') if type(x) == datetime else x[:10].replace('/','-'))  #Edita a variável no formato d - m - y
    
    data_pendente_novo = data_pendente.iloc[-count_row:] #cria uma tabela só dos novos (completo)
    
    #----------- Preparando respostas novas e comparando com antigo para ver respostas novas --------------------
    
    #*************Etapa igual das respostas_antigas
    tabela_respostas_novas_1 = pd.concat([pd.Series(row['Retorno'].split(',')) for _, row in data_pendente_novo.iterrows()]).reset_index(drop=True)
    tabela_respostas_novas_1 = tabela_respostas_novas_1.rename('Resposta')
    tabela_respostas_novas_1 = tabela_respostas_novas_1.drop_duplicates(keep='first')
    
    tabela_respostas_novas_2 = tabela_respostas_novas_1.apply(lambda x: 'Sinistro já cadastro' if re.search('já cadastrado',x) is not None else 
              ('Campos inválidos' if re.search('Campos inválidos',x) is not None else 
              ('Falha na execucao do callable ' if re.search('Falha na execucao do callable ',x) is not None else x)))
    tabela_respostas_novas_2 = tabela_respostas_novas_2.apply(lambda x: 'Campos inválidos' if re.compile(r'\d{8} inválido').search(x) is not None else x)
    tabela_respostas_novas_2 = tabela_respostas_novas_2.apply(lambda x: 'Campos inválidos' if re.compile(r'conteúdo " inválido').search(x) is not None else x)
    
    
    tabela_respostas_novas_3 = tabela_respostas_novas_2.apply(lambda x: x.replace("'","") if re.search("'",x) is not None else x)
    tabela_respostas_novas_3 = tabela_respostas_novas_3.apply(lambda x: x.replace("[","") if re.search("\[",x) is not None else x)
    tabela_respostas_novas_3 = tabela_respostas_novas_3.apply(lambda x: x.replace("]","") if re.search("\]",x) is not None else x)
    tabela_respostas_novas_3 = tabela_respostas_novas_3.replace("",'""')
    tabela_respostas_novas_3 = tabela_respostas_novas_3.apply(lambda x: x.strip())
    tabela_respostas_novas_3 = tabela_respostas_novas_3.apply(lambda x: 'Necessario informar ao menos uma cobertura para o objeto codigo' if re.search('Necessario informar ao menos uma cobertura para o objeto codigo',x) is not None else x)
    tabela_respostas_novas_3 = tabela_respostas_novas_3.apply(lambda x: 'Campo apolice intermediador invalido' if re.search('apolice intermediador',x) is not None else x)
    
    
    tabela_respostas_novas_3 = tabela_respostas_novas_3.drop_duplicates(keep='first')
    
    del tabela_respostas_novas_1, tabela_respostas_novas_2
    
    #------------------------------------------------
    
    # ---------- Inicio da comparação para verificar se tem respostas novas e avisar via email ------
    
    
    tabela_respostas_antigas_4 = set(tabela_respostas_antigas_3) #Colocando no type set, pois o mesmo por default deduplica os valores, mantendo respostas únicas.
    tabela_respostas_novas_4 = set(tabela_respostas_novas_3) #Colocando no type set, pois o mesmo por default deduplica os valores, mantendo respostas únicas.
    
    resposta_nova = tabela_respostas_novas_4.difference(tabela_respostas_antigas_4) #Compara as respostas dos arquivos novos e antigos para identificação de respostas novas
    resposta_nova = list(resposta_nova) #Retorna para lista, pois é mais volátil para trabalhar
    lista_arquivo_rpnova = [] #lista vazia para appendar respostas novas
    
    data_pendente_novo_arq = data_pendente_novo[['Arquivo no Micro','Retorno']].copy() #Copio a base mantendo as colunas 'Arquivo no Micro' e 'Retorno' para comparar as respostas novas retornando os arquivos que contém essa resposta para notificar para equipe de Data Lake
    
    #Realizando um data cleaning na variável de forma similar no tratamento das respostas dos arquivos antigos e novos
    data_pendente_novo_arq['Retorno'] = data_pendente_novo_arq['Retorno'].apply(lambda x: x.replace("'","") if re.search("'",x) is not None else x)
    data_pendente_novo_arq['Retorno'] = data_pendente_novo_arq['Retorno'].apply(lambda x: x.replace("[","") if re.search("\[",x) is not None else x)
    data_pendente_novo_arq['Retorno'] = data_pendente_novo_arq['Retorno'].apply(lambda x: x.replace("]","") if re.search("\]",x) is not None else x)
    data_pendente_novo_arq['Retorno'] = data_pendente_novo_arq['Retorno'].replace("",'""')
    data_pendente_novo_arq['Retorno'] = data_pendente_novo_arq['Retorno'].apply(lambda x: x.strip())
    
    for i in resposta_nova: #Iterar as respostas novas 
        for colval,arquivo in zip(data_pendente_novo_arq['Retorno'],data_pendente_novo_arq['Arquivo no Micro']): #Para cada resposta nova, iterar na coluna afim de identificar qual arquivo contém aquela resposta
            if re.search(i,colval) is not None: #Se a substring (resposta) está naquela string 
               lista_arquivo_rpnova.append([i,arquivo]) #Appenda o nome do arquivo e a resposta
            else: #caso não contém aquela substring (resposta), continua a iteração
               pass
    
    if not lista_arquivo_rpnova: #Se a lista for vazia, continua o código
        pass
    else: #Se não for vazio (teve resposta nova), envia um email mostrando o arquivo e resposta nova.
        Template =  r""" 
                    <hl> As respostas abaixo são novas e devem ser comunicadas com urgência para equipe de DataLake:</hl>
                    <p> %s <p>
                    <p> %s <p>
                    <p>Att,<p>""" % ("\n".join(map(str,lista_arquivo_rpnova[0])),"\n".join(map(str,lista_arquivo_rpnova[1]))) #template para criação do corpo de email, usando o join onde vai colocar o arquivo na parte superior, e inferior será a respota
        outlook = win32.Dispatch('outlook.application') #Usado para criar uma conexão com o Microsoft Outlook. O "Dispatch" cria uma instância de COM (Component Object Model) de um objeto specificado pelo argumento "outlook.application"
        mail= outlook.CreateItem(0) #O método 'CreateItem' cria uma instância do tipo de item especificado e retorna uma referência a ele, que o parâmetro 0 é mensagem de email.
        mail.To = "**********" #Parâmetro to - Colocar o email para o destinatário que receberá a mesma.
        mail.Subject = "Respostas novas - B3" #Parâmetro subject para colocar o título do email
        mail.htmlBody = Template #htmlBody parâmetro para retornar o template criado em HTML para aparecer no corpo do email
        inspector = mail.GetInspector #Retorna um objeto Inspector representando a janela do inspetor para o item recém-criado, que é uma interface para acessar as propriedades e métodos do item que está sendo inspecionado.
        inspector.Display() #O método do objeto Inspetor exibe a janela do inspetor, que permite ao usuário visualizar e editar o item.
        doc=inspector.WordEditor #A propriedade do objeto Inspector retorna o objeto WordEditor, que fornece acesso ao modelo de objeto de documento do Microsoft Word (DOM) do item do email. Isso permite que o usuário manipule programaticamente o conteúdo do item de email usando o Microsoft Word.
        selection=doc.Content #A propriedade do objeto WordEditor retorna o conteúdo do item de email como uma string. Isso permite que o usuário acesse e manipule programaticamente o conteúdo do item de email.
        
        mail.Send() #Enviar o email criado
    
    #---------------------------------------------------------
    
    
    #----------------- Preenchimento do campo "Pendência" para classificar se o arquivo está como pendente ou sem pendencia para reprocessamento via Insurconnect B3 --------------------
    
    lista_ok = ['EXECUCAO OK','EXECUÇÃO OK',' Nº da Apólice já cadastrada', ' Já existe um endosso registrado com o número informado', ' Identificador Movimento já cadastrado'] #Respostas de identificação de "sem pendência"
    
    
    data_pendente_novo['Pendência'] = data_pendente_novo['Pendência'].fillna('Vazio') #Substituindo os valores vazios pela string 'Vazio'
    
    column_list = data_pendente_novo['Retorno'].tolist() #Colocando o campo 'Retorno' em um objeto lista para melhor manipulação do dado
    lista_pendente = [] #Lista vazia para appendar a pendência
    lista_erro = [] #Lista vazia para appendar o erro que deve ser classificado como 'pendente', ou seja, diferente da 'lista_ok'
    
    #---- Nessa etapa foi utilizado uma lógica para identificar o erro por valor e eles estão como uma lista dentro de uma string em cada célula -----------
    
    for i in column_list: #iterando a lista 'Retorno' para identificação de cada resposta, se é erro ou nao
        variavel_imutavel = i #Mantendo a resposta imutável pós tratamento do i, pois necessitará para lógica
        if i[0] != "[" and i[0] != "'" and i[-1] != "]" and i[-1] != "'": #Se a primeira letra da string for diferente tanto de "[" e "'" e o último "]" e "'", vamos adicionar "['" e "]" para conseguir realizar a separação do retorno
            i = "['"+i+"']"
        elif i[0]=="[" and i[1] == "'" and i[-2]=="'" and i[-1]=="]": #Se ele estiver no inicio "['" e parte final "']"ele retornará ele mesmo
           i=i 
        elif i[0]=="[":#Se ele começar com o caracter "[" vai retornar ele mesmo 
            i = i
        elif i[0]=="'": #Se ele começar com o caracter "'" vai adicionar "["
            i = "["+i
        else: #Caso não possua as estruturas anterior, vai adicionar "['"
            i="['"+i
        
        
        if  variavel_imutavel[0] != "[" and variavel_imutavel [0] != "'" and variavel_imutavel [-1] != "]" and variavel_imutavel [-1] != "'": #Se
            i = i
        elif (i[0]!="[" and i[1]!="'") and i[-1]=="]" and i[-2]=="'" and i[-3]!="'" and "Campos inválidos" in i:
            i=i+"]"
        elif i[-1]=="'" and i[-2]!="'" and i[-2]!="]" and variavel_imutavel[0]!="[" and "Campos inválidos" in i:
            i = i+"]]"
        elif i[-1]!="'" and i[-1]!="]":
            i = i+"]']"
        else:
            i = i
    
        words = ast.literal_eval(i) #Ele basicamente vai interpretar uma string como valor literal de outro objeto Python como lista, propria string, tupla. Isso é para "transformar" a string como estrutura lista e separar os erros
        
        ident_pendente = " ".join([word for word in words if word not in lista_ok and re.search('já cadastrado',word) is None]) #Lista compreensiva, está pegando o valor que "transformado" em lista, verifica dentro dessa lista de resposta para aquele arquivo se um deles é erro, se for, faz um join em uma string vazia, caso não, não apenda
        lista_erro.append(ident_pendente) #Appenda as respostas que são erros
        if ident_pendente =="": #Se não tiver erro, retorna "Sem pendencia"
            lista_pendente.append('SEM PENDÊNCIA')
        else: #Se tiver algum erro, retorna pendente
            lista_pendente.append('PENDENTE')
            
    
    #--------------- Substituição da coluna de classificação de pendência pela coluna atualizada com status  -----------------
    
    data_pendente_novo = data_pendente_novo.reset_index(drop = True) #Resetando o index
    
    mySeries = pd.Series(lista_pendente) #Colocando a lista como Series pandas para substituir a coluna "Pendência" atualizada no Dataframe
    data_pendente_novo['Pendência'] = mySeries.values #Preenchendo a coluna "Pendência" do arquivo novo.
    
    data_pendente_final = data_pendente.copy() #Copio a base para trabalhar separado do Dataframe original
    
    data_pendente_final.iloc[-count_row:] = data_pendente_novo.values #Pegando os valores novos, atualizando com pendência, e coloco na base histórica
    
    data_teste = data_pendente_final.copy()  #Copio a base para trabalhar separado do Dataframe original
    
    data_teste['nome_arquivo_dph'] = data_teste['Arquivo no Micro'].apply(lambda x: x[0:22]+"DPH" #if re.search('REP1', x) is not None  else  #Crio uma coluna onde vai manter o nome dos arquivos sem ramo e diferenciando LEG, DPH e o "normal"
                                  # (x[0:22]+"REP2" if re.search('REP2', x) is not None else
                                  # (x[0:22]+"REP3" if re.search('REP3', x) is not None else 
                                  # (x[0:22]+"REP4" if re.search('REP4', x) is not None else
                                  # (x[0:22]+"REP5" if re.search('REP5', x) is not None else
                                  # (x[0:22]+"DIVAPO" if re.search("DIVAPO", x) is not None else
                                  # (x[0:22]+"REENV" if re.search("REENV", x) is not None   else
                                  # (x[0:22]+"DIVEND" if re.search("DIVEND", x) is not None else
                                  # (x[0:22]+"RPDT" if re.search("RPDT", x) is not None     else
                                  # (x[0:22]+"LEG90" if re.search("LEG90", x) is not None   else
                                  # (x[0:22]+"LEG92" if re.search("LEG92", x) is not None   else
                                  # (x[0:22]+"LEG89" if re.search("LEG89", x) is not None   else
                                  # (x[0:22]+"LEG102" if re.search("LEG102", x) is not None else
                                  # (x[0:22]+"REPV2" if re.search("REPV2", x) is not None   else
                                  # (x[0:22]+"COSSEGRESSEG_REP" if  re.search("COSSEGRESSEG_REP", x) is not None else
                                  # (x[0:22]+"COSSEGRESSEG" if re.search("COSSEGRESSEG", x) is not None else
                                  # (x[0:22]+"DPH_REP" if re.search("DPH_REP", x) is not None else
                                  #(x[0:22]+"DPH"
                                  if re.search("DPH", x) is not None else
                                  (x[0:22]+"LEG" if re.search("LEG", x) is not None else                             
                                  # (x[0:22]+"TESTE_V1" if re.search("TESTE_V1", x) is not None else
                                  # (x[0:22]+"TESTE_V1" if re.search("TESTE_V01", x) is not None else
                                  # (x[0:22]+"TESTE_V2" if re.search("TESTE_V2", x) is not None else
                                  # (x[0:22]+"TESTE_V2" if re.search("TESTE_V02", x) is not None else
                                  # (x[0:22]+"TESTE_V3" if re.search("TESTE_V3", x) is not None else
                                  # (x[0:22]+"TESTE_V3" if re.search("TESTE_V03", x) is not None else
                                  # (x[0:22]+"REP" if re.search('REP', x) is not None else   
                                   x[0:22]))#))))))))))))))))))))))))) 
    
    #---------------- Processo de identificação do ramo do arquivo para agrupar o arquivo por ramo----------------------
    
    data_arq_no_micro = data_teste['Arquivo no Micro'].tolist() #Colocando o campo 'Arquivo no Micro' em um objeto lista para melhor manipulação do dado
    lista_ramo = [] #Lista vazia para appendar os ramos dos arquivos e concatenar com o "Arquivo no Micro"
    
    for i in data_arq_no_micro:#iterando a "data_arq_no_micro" para retirar o ramo de cada arquivo e assim separar o arquivo por ramo
        try:
            if re.sub("[^0-9]", "", i.partition("_")[-1])[0]!="1": #Vai utilizar um regex para criar um padrão, onde vai pegar números [0-9], a string que ele pegará 
                                                                   #do valor particionado pela função partition, onde vai separar a string pelo caracter "_" pegando sempre 
                                                                   #O primeiro valor ao usar [-1], e compara, se for diferente de 1, então só tem apenas um caracter do ramo
                ramo_unitario=re.sub("[^0-9]", "", i.partition("_")[-1])[0] #Colocando o ramo na variável
                lista_ramo.append(ramo_unitario) #Appendando/Salvando o ramo em uma lista
            else:
                try:#Caso seja 1, vai tentar pegar o outro número para caso seja uma dezena (10,11,12,...), se der erro, o ramo é "1"
                    ramo_string = i.partition("_")[-1] #string com o ramo
                    ramo_dezena=re.sub("[^0-9]", "", ramo_string)[0]+re.sub("[^0-9]", "", ramo_string)[1] #Aqui extrai a dezena e a unidade do ramo (10,11,...)
                    lista_ramo.append(ramo_dezena) #Appendando/Salvando o ramo em uma lista
                except: #É o caso que é uma unidade igual "1" o ramo
                    ramo_string = i.partition("_")[-1] #string com o ramo
                    ramo_unidade_1=re.sub("[^0-9]", "", ramo_string)[0] #Extrai o ramo "1"
                    lista_ramo.append(ramo_unidade_1) #Appendando/Salvando o ramo em uma lista
                    
            
        except: #se ele não encontrar ramo, deixará vazio
            lista_ramo.append("") #Appendando vazio
    
    ramo_serie = pd.Series(lista_ramo) #Colocando a lista como Series pandas para "somar" o ramo com a coluna "nome_arquivo_dph" no Dataframe
    data_atualiza_ramo = data_teste.copy() #Copio a base para trabalhar separado do Dataframe original
    data_atualiza_ramo['nome_arquivo_dph'] = data_atualiza_ramo['nome_arquivo_dph'] + ramo_serie.values #"Somando" o ramo no 'nome_arquivo_dph' para separar o arquivo pelo ramo
    
    #----------------- Criando a lógica de pendência - caso o arquivo reprocessado seja pendente,
    
    data_atualiza_ramo['flag'] = np.where(data_atualiza_ramo['Pendência']== 'PENDENTE', 1, 0) #Criando a variável 'flag' para caso for 'Pendente' é 1, caso seja 'Sem pendencia', 0
    data_atualiza_ramo = data_atualiza_ramo.drop_duplicates(keep='first') #Deduplico a base para não ter repetição
    
    data_agrupada_count_max = data_atualiza_ramo.groupby(['Data de Recebimento','nome_arquivo_dph'])['flag'].agg(['max','count']) #Base agrupada pegando o flag max, ou seja, se é pendente (1) e a contagem de flag
                                                                                                                                  #Para verificar se o arquivo tem partição (>1)
    
    data_agrupada = data_agrupada_count_max.reset_index() #Reseta o index para não ficar bagunçado e duplicado o index
    
    data_agrupada['Data de Recebimento'] = pd.to_datetime(data_agrupada['Data de Recebimento']) #Modificando o type da coluna "Data de Recebimento" de string para datetime64
    
    data_agrupada = data_agrupada.sort_values(by=['nome_arquivo_dph','Data de Recebimento'],ascending = [True, True]) #Ordena pelo 'nome_arquivo_dph' e depois 'Data de Recebimento' de forma crescente
    data_agrupada = data_agrupada.reset_index(drop=True) #Reseta o index para não ficar bagunçado e duplicado o index
    data_agrupada = data_agrupada.drop_duplicates(keep='last') #Deduplica o agrupamento mantendo o último
    
    data_agrupada_1 = data_agrupada.drop_duplicates(subset=['nome_arquivo_dph'], keep='last') #Deduplica por "nome_arquivo_dph" mantendo o último, ou seja, data mais recente
    data_agrupada_2 = data_agrupada_1.reset_index(drop=False) #Reseta o index para não ficar bagunçado e duplicado o index
    
    
    # ---------- Base original sem agrupamento para cruzamento com base agrupada para modificação de pendência -----------
    
    data_atualiza_ramo['Data de Recebimento'] = pd.to_datetime(data_atualiza_ramo['Data de Recebimento']) #Modificando o type da coluna "Data de Recebimento" de string para datetime64
    data_atualiza_ramo = data_atualiza_ramo.sort_values(by='Data de Recebimento') #Ordena pela 'Data de Recebimento'
    data_atualiza_ramo=data_atualiza_ramo.reset_index(drop=True) #Reseta o index para não ficar bagunçado e duplicado o index
    
    # ---------- Cruzando a base original com a agrupada para pegar o 'max' e 'count' --------------
    
    base_pendente_final = pd.merge(data_atualiza_ramo,data_agrupada_2[['nome_arquivo_dph','max','count']],on=["nome_arquivo_dph"],how="left") #Faz um left join para pegar o max e arquivo_dph para identificar se o máximo daquele arquivo é 'pendente'
                                                                                                                                              #e também quantidade de arquivo (se tem partição)
    
    #---------- Modificando o status para pendente, replicando os arquivo reprocessados que na última data estava sem pendência para os demais ----------------------
    base_pendente_final['Pendência_1'] = np.where((base_pendente_final['max'] == 0 ) & (base_pendente_final['count'] > 1), "SEM PENDÊNCIA",  # Se o máximo daquele arquivo é 0 (sem pendência) e com partição (>1) retorna 'Sem pendência'
                                         np.where((base_pendente_final['max'] == 1 ) & (base_pendente_final['count'] > 1), base_pendente_final['Pendência'], # Se o máximo daquele arquivo é 1 (pendente) e com partição (>1) retorna o mesmo valor que está (Pendente ou Sem pendencia)
                                         np.where((base_pendente_final['max'] == 0 ) & (base_pendente_final['count'] == 1),"SEM PENDÊNCIA", # Se o máximo daquele arquivo é 0 (sem pendência) e sem partição retorna 'Sem pendência'
                                         np.where((base_pendente_final['max'] == 1 ) & (base_pendente_final['count'] == 1), base_pendente_final['Pendência'], # Se o máximo daquele arquivo é 1 e sem partição retorna ele mesmo (que vai ser pendente)
                                         base_pendente_final['Pendência'])))) #Caso não seja nenhuma dessas condições, vai retornar ele mesmo ('Pendente' ou 'Sem pendência')
    
    
    #----------- Cruzamento a nova coluna 'Pendência-1' com o arquivo original DESORDENADO -------------------------------------
    
    data_pendente['Data de Recebimento'] = pd.to_datetime(data_pendente['Data de Recebimento']) #Modificando o type da coluna "Data de Recebimento" de string para datetime64
    
    base_completa = pd.merge(data_pendente,base_pendente_final[['Data de Recebimento','Arquivo no Micro','ID','Pendência_1']],on=['Data de Recebimento','Arquivo no Micro','ID'],how="left") #Left join da pendência alterada com o arquivo original (DESORDENADO)
    
    pendencia_corrigida = base_completa['Pendência_1'] #Dataframe com apenas a coluna 'Pendência_1' para substituir a pendência original (devido ao critério acima)
    
    #-------------- Mesma etapa do processo de escrita no Excel feito anteriormente ---------
    book = load_workbook(path, keep_vba=True) 
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    pendencia_corrigida.to_excel(writer, sheet_name = 'Lista arquivos', startrow =1 ,startcol= 13, index = False, header=False) #Aqui vai subescrever a coluna inteira de 'pendência' colocando o atualizado
    
    
    writer.save()
    
    #--------------- Mesma etapa de ativar a macro apenas abrindo e fechando o Excel---------------
    
    xl_app = xs.App(visible=False, add_book=False)
    wb = xl_app.books.open(path)
    
    wb.save()
    wb.close()
    
    #--------------- Etapa de verificação de processamento dos arquivos: 1) Se foi enviado para B3 mas não está no Insurconnect 2) Foi processado, apareceu no Insurconnect, mas não está no arquivo de Retorno -----------------
    
    
    #---------------- -------1)Se foi enviado para B3 mas não está no Insurconnect ---------------------------
    messages_proc_sro = inbox.Items.Restrict("@SQL=(urn:schemas:httpmail:subject LIKE '%Processamento SRO%')") #Objeto com as mensagens do INBOX filtrados pelo título com a escrita "Processsamento SRO"
    
    dia_de_hoje = date.today().strftime("%A") #Colocando o dia da semana da data do dia em uma variável (Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday)
    
    lista_arquivos_proc  = [] #Lista vazio para appendar os arquivos enviados para B3 estão no corpo do email enviado automaticamente após envio
    if dia_de_hoje == "Monday": #Se for Segunda, vamos pegar Segunda, Domingo, Sábado e Sexta
        lista_proc_sro = list(messages_proc_sro) #Colocando os emails em uma lista Python para melhor manipulação do dado
        for i in (1,2,3,4): #iterando para o range de 1-4 que pegará os arquivos dos 4 dias
            ultimo_email_proc_sro = lista_proc_sro[-i] #Colocando o email da iteracao em uma variável
            view_proc_sro = ultimo_email_proc_sro.HTMLBody #Pegar o corpo de email que de fato é onde contém o texto dos arquivos
            string_arquivos_proc_sro = view_proc_sro[101:-26] #A partir da posição da string, capturamos a substring que contém o nome dos arquivos
            lista_arquivos_proc_sro = string_arquivos_proc_sro.split('<br>') #Retirando a substring '<br>' que não servirá, e gera como lista
            arquivos_proc_sro_atual = [elemento for elemento in lista_arquivos_proc_sro if elemento.strip() != ""] #Retirando valor vazio da lista que é gerado
            lista_arquivos_proc.append(arquivos_proc_sro_atual) #Appendando/Salvando os arquivos enviados de acordo com o email
        listaUnica = list(itertools.chain(*lista_arquivos_proc)) #Como termina sendo uma lista dentro de uma lista, usa o itertools que coloca os items das listas em uma lista única
    else: #Caso não seja segunda, vai iterar apenas com o dia atual e o anterior, 2 dias, executando o mesmo processo anterior    
        lista_proc_sro = list(messages_proc_sro)
        for i in (1,2):
            ultimo_email_proc_sro = lista_proc_sro[-i]
            view_proc_sro = ultimo_email_proc_sro.HTMLBody
            string_arquivos_proc_sro = view_proc_sro[101:-26]
            lista_arquivos_proc_sro = string_arquivos_proc_sro.split('<br>')
            arquivos_proc_sro_atual = [elemento for elemento in lista_arquivos_proc_sro if elemento.strip() != ""]
            lista_arquivos_proc.append(arquivos_proc_sro_atual)
        listaUnica = list(itertools.chain(*lista_arquivos_proc))
    
    
    nao_encontrado_b3 = set(listaUnica).difference(set(lista_arquivos_novos_b3)) #Verificando se tem arquivos diferentes entre arquivos B3 InsurConnect e arquivos demonstrados no email automático de enviados
    
    
    if not nao_encontrado_b3: #Se está vazio, ou seja, todos os arquivos foram processados na B3 Insurconnect, continua o programa
        pass
    else: #Se tem arquivo, ou seja, não foi processado, encaminha email informando
    
        #Processo demonstrando anteriormente
        Template =  r""" 
                    <hl> Os arquivos abaixo não foram processados pela B3 - InsurConnect e deve ser comunicado com urgência </hl>
                    <p> %s <p>
                    <p>Att,<p>""" % ("\n".join(nao_encontrado_b3))
        outlook = win32.Dispatch('outlook.application')
        mail= outlook.CreateItem(0)
        mail.To = "*************"
        mail.Subject = "Arquivos não processados - B3"
        mail.htmlBody = Template
        inspector = mail.GetInspector
        inspector.Display()
        doc=inspector.WordEditor
        selection=doc.Content
        
        mail.Send()
    
    #---------------- -------2) Foi processado, apareceu no Insurconnect, mas não está no arquivo de Retorno ---------------------------
    
    pathRG = [r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******'] #Lista com os caminhos dos arquivos recebidos da RIOGRANDE
    pathIS = [r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******',r'//ISMTZ*******'] #Lista com os caminhos dos arquivos recebidos da ICATU
    
    lista_arq_recebidos=[] #Lista vazio para appendar os arquivos recebidos pela B3 pós InsurConnect
    for (i,j) in zip(pathRG,pathIS): #Iterando ao mesmo tempo (zip) tanto na Riogrande quanto Icatu
        all_files_RG = glob.glob(i + "/*.csv.*" , recursive = True) #Cria um arquivo do tipo lista de acordo com caminho pegando todos os mesmos que terminam com .csv, sendo esses valores de caminho + nome do arquivo da RioGrande
        all_files_IS = glob.glob(j + "/*.CSV.*", recursive = True) #Cria um arquivo do tipo lista de acordo com caminho pegando todos os mesmos que terminam com .CSV, sendo esses valores de caminho + nome do arquivo da Icatu
        names_RG = [os.path.basename(x) for x in all_files_RG] #Usando a função para extrair o diretorio, mantendo só o nome do arquivo RioGrande
        names_IS = [os.path.basename(x) for x in all_files_IS] #Usando a função para extrair o diretorio, mantendo só o nome do arquivo Icatu
        all_files = names_RG + names_IS #Soma em uma única lista todos os arquivos
        lista_arq_recebidos.append(all_files) #Appendando/Salvando todos os arquivos em uma lista
        
    lista_arq_recebidos = list(itertools.chain(*lista_arq_recebidos)) #Como termina sendo uma lista dentro de uma lista, usa o itertools que coloca os items das listas em uma lista única
        
    lista_arq_recebidos_1 = [w[:-8] for w in lista_arq_recebidos] #Para cada item da lista, vai manter só o nome, retirando o ID do arquivo   
    lista_arq_recebidos_2 = [w[-7::] for w in lista_arq_recebidos] #Para cada item da lista, vai manter só o ID, retirando o nome do arquivo    
    
    data_tuples = list(zip(lista_arq_recebidos_1,lista_arq_recebidos_2)) #Coloca no formato tupla para transformar em um Dataframe
    recebidos_b3 = pd.DataFrame(data_tuples, columns=['Arquivo no Micro','ID']) #Transforma a tupla em Dataframe nomeando o nome de cada coluna
    
    nao_encontrado_pasta = set(lista_arquivos_novos_b3).difference(set(lista_arq_recebidos_1)) #Verificando se tem arquivos diferentes entre arquivos B3 InsurConnect e arquivos que estão nas pastas dos recebidos pós InsurConnect
    
    if not nao_encontrado_pasta: #Se está vazio, ou seja, todos os arquivos estão nas pastas recebidos
        pass
    else: #Se tem arquivo, ou seja, não está nos arquivos recebidos, encaminha email informando
        
        #Processo demonstrando anteriormente
        Template =  r""" 
                    <hl> Os arquivos abaixo estão no lançamento via InsurConnect, mas não se apresentam na pasta 'recebidos': </hl>
                    <p> %s <p>
                    <p>Att,<p>""" % ("\n".join(nao_encontrado_pasta))
        outlook = win32.Dispatch('outlook.application')
        mail= outlook.CreateItem(0)
        mail.To = "*******"
        mail.Subject = "Arquivos não recebidos - B3"
        mail.htmlBody = Template
        inspector = mail.GetInspector
        inspector.Display()
        doc=inspector.WordEditor
        selection=doc.Content
        
        mail.Send()
        
    #----------------- Retirando duplicada com VBA (rodando macro da Planilha Insurconnect) ------------------------
    
    excel = win32.Dispatch("Excel.Application") #Crie uma conexão com uma instância do Microsoft Excel em execução em um sistema operacional Windows. Esse código cria um objeto COM que representa o aplicativo Excel, permitindo que um programa Python interaja com o Excel usando a interface de automação COM.
    excel.Visible = False #A propriedade do objeto Excel determina se a janela do aplicativo Excel está visível ou oculta. Por padrão, a propriedade Visible está configurada para True, o que significa que a janela do aplicativo Excel está visível.
    workbook = excel.Workbooks.Open(path) #Abrirá a worbook que está no caminho 'path'
    worksheet = workbook.Worksheets("Lista arquivos") #Cria uma variável que é a sheet específica selecionada
    
    excel.Run("Módulo1.RemoveDuplicateRows") #Roda a macro específica para remover linhas duplicadas
    
    workbook.Save() #Salva a worbook que foi atualizada pela macro
    workbook.Close() #Fecha a worbook
    
    excel.Quit() #Sai do Excel
    
    texto_funcionando.config(text="O programa foi finalizado!") #Altera o texto da Label criada na GUI do Tkinter após finalizar o programa

#---------------------------------------------------------------------

#------------ Criando uma função que executa um thread para permitir que o programa continue executando outro código enquanto espera que as operações de E/S sejam concluídas em segundo plano.-----------------

def button_callback(d_i,m_i,a_i,d_f,m_f,a_f):
    t = threading.Thread(target=monitoramento_automatico, args=(d_i,m_i,a_i,d_f,m_f,a_f)) #Isso é útil para executar tarefas de execução longa em segundo plano enquanto o programa principal continua respondendo à entrada do usuário ou executando outras tarefas.
    t.start()
    

#----------- Desenvolvimento do GUI (Interface gráfica para o usuário) ----------------------

root = tk.Tk() #Variável da janela que será inserido os Widgets 
root.geometry("400x150") #Dimensão da janela
root.title("Automatização do monitoramento SRO") #Título da janela 

# Criação das labels
periodo = ttk.Label(root, text="Período") #Label com o texto Período
periodo_inicial = ttk.Label(root, text="Inicial") #Label com o texto Inicial
periodo_final = ttk.Label(root, text="Final") #Labeel com o texto final
ex_inicial = ttk.Label(root, text="Ex: 5 - 2 - 2023") #Label com o texto informativo do exemplo para inserir no Entry inicial
ex_final = ttk.Label(root, text="Ex: 8 - 2 - 2023") #Label com o texto informativo do exemplo para inserir no Entry final


traco1_label = ttk.Label(root, text="-") #Label com o traço para separar as Entry
traco2_label = ttk.Label(root, text="-") #Label com o traço para separar as Entry
traco3_label = ttk.Label(root, text="-") #Label com o traço para separar as Entry
traco4_label = ttk.Label(root, text="-") #Label com o traço para separar as Entry

texto_funcionando = tk.Label(root, text="") #Label vazia para alteração de status quando começar e finalizar o programa


# Criando uma Entry (input) para cada parte do período (data)

day_box = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do dia inicial com largura específica 6 e texto centralizado ao inputar
month_box = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do mês inicial com largura específica 6 e texto centralizado ao inputar
year_box = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do ano inicial com largura específica 6 e texto centralizado ao inputar
day_box_final = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do dia final com largura específica 6 e texto centralizado ao inputar
month_box_final = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do mês final com largura específica 6 e texto centralizado ao inputar
year_box_final = ttk.Entry(root, width=6, justify='center') #Criando uma Entry do ano final com largura específica 6 e texto centralizado ao inputar

# Criando o botão para rodar a função 

botao = tk.Button(root, text = 'Run', command= lambda: button_callback(day_box.get(),month_box.get(),year_box.get(),day_box_final.get(),month_box_final.get(),year_box_final.get())) # Criando o botão com texto "run" para rodar a função "monitoramento_automatico" em segundo plano na thread dentro de button_calback que terá os parametros inputados a partir dos Entry



# Colocando as labels, botões e Entry na janela 

periodo.grid(row=0, column=0, padx=10, pady=10) #Colocando a label na linha 0 e coluna 0, com espaçamento horizontal 10 e vertical 10
periodo_inicial.grid(row=1, column=0, padx=10, pady=5) #Colocando a label na linha 1 e coluna 0, com espaçamento horizontal 10 e vertical 5
periodo_final.grid(row=2, column=0, padx=10, pady=5) #Colocando a label na linha 2 e coluna 0, com espaçamento horizontal 10 e vertical 5

day_box.grid(row=1, column=1, padx=10, pady=5) #Colocando a Entry na linha 1 e coluna 1, com espaçamento horizontal 10 e vertical 5
month_box.grid(row=1, column=3, padx=10, pady=5) #Colocando a Entry na linha 1 e coluna 3, com espaçamento horizontal 10 e vertical 5
year_box.grid(row=1, column=5, padx=10, pady=5) #Colocando a Entry na linha 1 e coluna 5, com espaçamento horizontal 10 e vertical 5

day_box_final.grid(row=2, column=1, padx=10, pady=5) #Colocando a Entry na linha 2 e coluna 1, com espaçamento horizontal 10 e vertical 5
month_box_final.grid(row=2, column=3, padx=10, pady=5) #Colocando a Entry na linha 2 e coluna 3, com espaçamento horizontal 10 e vertical 5
year_box_final.grid(row=2, column=5, padx=10, pady=5) #Colocando a Entry na linha 2 e coluna 5, com espaçamento horizontal 10 e vertical 5

traco1_label.grid(row=1, column=2, padx=2, pady=5) #Colocando a label na linha 1 e coluna 2, com espaçamento horizontal 2 e vertical 5
traco2_label.grid(row=1, column=4, padx=2, pady=5) #Colocando a label na linha 1 e coluna 4, com espaçamento horizontal 2 e vertical 5

traco3_label.grid(row=2, column=2, padx=2, pady=5) #Colocando a label na linha 2 e coluna 2, com espaçamento horizontal 2 e vertical 5
traco4_label.grid(row=2, column=4, padx=2, pady=5) #Colocando a label na linha 2 e coluna 4, com espaçamento horizontal 2 e vertical 5

ex_inicial.grid(row=1, column=6, padx=2, pady=5) #Colocando a label na linha 1 e coluna 6, com espaçamento horizontal 2 e vertical 5
ex_final.grid(row=2, column=6, padx=2, pady=5) #Colocando a label na linha 2 e coluna 6, com espaçamento horizontal 2 e vertical 5

botao.grid(row=3, column=0, padx=2, pady=5) #Colocando a label na linha 3 e coluna 0, com espaçamento horizontal 2 e vertical 5

texto_funcionando.grid(row=3, column=1, columnspan=6,padx=2, pady=5) #Colocando a label na linha 3 e coluna 1, com espaçamento horizontal 2 e vertical 5

root.mainloop() # Ele inicia o loop de eventos do aplicativo e o programa permanece nesse loop até que o usuário feche a janela ou saia do aplicativo.